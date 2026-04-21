"""
Получить API Key
pip install openai
"""

import os
import re
import torch
import easyocr  # <--- Замена Paddle
from spellchecker import SpellChecker
from tqdm.auto import tqdm
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import time

from collections import Counter
import pandas as pd

from openai import OpenAI

from envs import OPENROUTER_TOKEN, HF_TOKEN, DEEPSEEK_API_KEY, PT_API_KEY

# --- ВАШИ ПУТИ (СОХРАНЕНЫ) ---
BASE_PATH = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo"
# train_path = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo\train\train_set"
full_set_path = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo\full_pics_lib"
set_path_VtM = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo\vampire_lib"
OUTPUT_FILE = os.path.join(BASE_PATH, "submission_final.csv")
OUTPUT_FILE_TINY = os.path.join(BASE_PATH, "submission_final_tiny.csv")
OUTPUT_FILE_FULL = os.path.join(BASE_PATH, "submission_final_full.csv")
# CSV_FILE = os.path.join(BASE_PATH, "translated_csv_dictionary.csv")
EXCEL_FILE = os.path.join(BASE_PATH, "MyVocabulary.xlsx")
EXCEL_FILE_TINY = os.path.join(BASE_PATH, "MyVocabulary_tiny.xlsx")

""" VtM """
OUTPUT_FILE_VtM1 = os.path.join(BASE_PATH, "submission_final_VtM_Copy_1.csv")
OUTPUT_FILE_VtM2 = os.path.join(BASE_PATH, "submission_final_VtM_Copy_2.csv")
OUTPUT_FILE_VtM3 = os.path.join(BASE_PATH, "submission_final_VtM_Copy_3.csv")
OUTPUT_FILE_VtM4 = os.path.join(BASE_PATH, "submission_final_VtM_Copy_4.csv")
OUTPUT_FILE_VtM5 = os.path.join(BASE_PATH, "submission_final_VtM_Copy_5.csv")
EXCEL_FILE_VtM1 = os.path.join(BASE_PATH, "MyVocabulary_VtM_1.xlsx")
EXCEL_FILE_VtM2 = os.path.join(BASE_PATH, "MyVocabulary_VtM_2.xlsx")
EXCEL_FILE_VtM3 = os.path.join(BASE_PATH, "MyVocabulary_VtM_3.xlsx")
EXCEL_FILE_VtM4 = os.path.join(BASE_PATH, "MyVocabulary_VtM_4.xlsx")
EXCEL_FILE_VtM5 = os.path.join(BASE_PATH, "MyVocabulary_VtM_5.xlsx")

class BatchEnglishExtractor:
    def __init__(self):
        print("--- Инициализация ---")

        # Используем PyTorch для проверки GPU (так как он у вас точно работает)
        self.use_gpu = torch.cuda.is_available()
        if self.use_gpu:
            print(f"[✓] GPU обнаружена: {torch.cuda.get_device_name(0)}")
        else:
            print("[x] GPU не найдена. Работаю на CPU.")

        # Инициализация EasyOCR
        # verbose=False убирает лишний шум в консоли
        self.reader = easyocr.Reader(['en'], gpu=self.use_gpu, verbose=False)

        # Инициализация словаря
        self.spell = SpellChecker()

    def _is_valid_english(self, text):
        """Фильтр 'Двойной замок': Unicode + Словарь"""
        clean_text = re.sub(r'[^\w\s]', '', text).strip()

        # Пропускаем мусор (цифры) и короткие слова
        if not clean_text or clean_text.isdigit() or len(clean_text) < 2:
            return False

        # Шаг Б: Unicode-ловушка (Кириллица)
        if re.search(r'[\u0400-\u04ff]', text):
            return False

        # Шаг В: Словарная проверка
        words_to_check = [clean_text, clean_text.lower()]
        # known() возвращает список найденных слов. Если список не пуст -> True
        return bool(self.spell.known(words_to_check))

    def process_folder(self, folder_path, output_file):
        valid_extensions = ('.jpg', '.jpeg', '.png', '.bmp')

        if not os.path.exists(folder_path):
            print(f"ОШИБКА: Папка {folder_path} не найдена.")
            return

        files = [f for f in os.listdir(folder_path) if f.lower().endswith(valid_extensions)]
        print(f"Найдено изображений: {len(files)}")

        with open(output_file, "w", encoding="utf-8") as f_out:
            # Используем tqdm для красивой полоски прогресса
            for filename in tqdm(files, desc="Обработка"):
                img_path = os.path.join(folder_path, filename)
                try:
                    # EasyOCR возвращает список кортежей: (bbox, text, confidence)
                    result = self.reader.readtext(img_path)
                    if not result:
                        continue

                    # Сбор данных
                    page_words = []
                    for (bbox, text, prob) in result:
                        # EasyOCR bbox формат: [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
                        # Берем y1 (верхний левый угол) для сортировки
                        y_coord = bbox[0][1]
                        page_words.append({"text": text, "y": y_coord})

                    # Сортировка сверху вниз
                    page_words.sort(key=lambda x: x["y"])

                    # Фильтрация и запись
                    for item in page_words:
                        word = item["text"]
                        if self._is_valid_english(word):
                            f_out.write(word + "\n")

                except Exception as e:
                    print(f"\nОшибка в файле {filename}: {e}")

        print(f"\nГотово! Все слова сохранены в: {output_file}")

    def analyze_frequency(self, input_file, top_n=10):
        """Считает частотность и выводит статистику в консоль."""
        with open(input_file, 'r', encoding='utf-8') as file:
            words = [line.strip() for line in file if line.strip()]

        # Считаем частотность
        word_counts = Counter(words)

        # Сортируем по популярности (от самых частых к редким)
        common_words = word_counts.most_common()

        print(f"--- Статистика файла: {input_file} ---")
        print(f"Всего слов (с повторами): {len(words)}")
        print(f"Уникальных слов: {len(word_counts)}")
        print(f"\nТоп-{top_n} частых слов:")

        for word, count in common_words[:top_n]:
            print(f"{word}: {count}")
        return common_words


    def remove_top_noise(self, input_file, noise_count=7):
        """Находит топ N самых частых слов и удаляет их из файла полностью."""
        with open(input_file, 'r', encoding='utf-8') as file:
            all_words = [line.strip() for line in file if line.strip()]

        # Находим, какие именно слова являются шумом
        word_counts = Counter(all_words)
        noise_words = set(word for word, count in word_counts.most_common(noise_count))
        # Когда делаем if word not in noise_words, Python в случае со списком каждый раз пробегает по нему.
        # В случае с set (множеством) поиск происходит мгновенно. На 600 словах разницы нет, но на 10 000+
        # это станет заметно.

        print(f"Удаляем шум: {', '.join(noise_words)}")

        # Оставляем только те слова, которых нет в списке шума
        filtered_words = [word for word in all_words if word not in noise_words]

        with open(input_file, 'w', encoding='utf-8') as f:
            for word in filtered_words:
                f.write(f"{word}\n")

        print(f"Файл {input_file} очищен от шума")
        return filtered_words  # Опционально


    def remove_duplicates(self, input_file):
        """Читает файл и возвращает список уникальных слов в порядке их появления."""
        with open(input_file, 'r', encoding='utf-8') as file:
            words = [line.strip() for line in file if line.strip()]

        # Использование dict.fromkeys сохраняет порядок первого появления слова
        unique_words = list(dict.fromkeys(words))

        with open(input_file, 'w', encoding='utf-8') as f:
            for word in unique_words:
                f.write(f"{word}\n")

        print(f"Файл {input_file} успешно обновлен.")

        return unique_words  # Опционально


class EnglishContextTranslator:
    def __init__(self, api_key):
        print("--- Инициализация ---")
        # Создаем клиент один раз при создании экземпляра класса
        self.api_key = api_key
        self.client = OpenAI(api_key=api_key,base_url="https://api-llm.ml.ptsecurity.ru/v1", timeout=120.0)
        self.model = "Qwen/Qwen35-397B-A17B-FP8"
       #  self.model = "openai/gpt-oss-120b"
       # self.model = "leon-se/gemma-3-27b-it-FP8-Dynamic"
       #  self.model = "Qwen/Qwen3-Coder-30B-A3B-Instruct"
       # self.model = "zai-org/GLM-47-Flash"

        # Используем PyTorch для проверки GPU (так как он у вас точно работает)
        self.use_gpu = torch.cuda.is_available()
        if self.use_gpu:
            print(f"[✓] GPU обнаружена: {torch.cuda.get_device_name(0)}")
        else:
            print("[x] GPU не найдена. Работаю на CPU.")


    def translate_word(self, word, system_prompt, temperature=0.1):
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"Translate: {word}"},
                ],
                stream=False,
                temperature=0.1  # Оставляем нашу строгость
            )
            if not response or not response.choices:
                return "Parsing error"

            answer = response.choices[0].message.content

            if not answer or answer.strip() == "":
                print(f"  [!] Пустой текст ответа для '{word}'.")
                return "Parsing error"
            return answer.strip()

        except Exception as e:
            if "429" in str(e):
                print("🛑 Лимит модели. Ждем 15 сек...")
                time.sleep(15)
                return self.translate_word(word, system_prompt)  # Повторная попытка
            else:
                print(f"API Error on word {word}: {e}")
                return "Parsing error"

    def judge_adequacy(self, word, translation, judge_system_prompt):
        """Модель оценивает свою (или чужую) работу через API"""
        audit_content = f"Word: {word}\nTranslation: {translation}\nRate adequacy (1-5):"
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": judge_system_prompt},
                    {"role": "user", "content": audit_content},
                ],
                stream=False,
                temperature=0.1
            )
            if not response or not response.choices:
                return 0

            answer = response.choices[0].message.content

            if not answer or answer.strip() == "":
                return 0

            match = re.search(r'\d', answer)
            return int(match.group()) if match else 0

        except Exception as e:
            if "429" in str(e):
                print(f"🛑 Лимит модели для Judge на '{word}'. Ждем 15 сек...")
                time.sleep(15)
                return self.judge_adequacy(word, translation, judge_system_prompt)
            else:
                print(f"[!] Ошибка судьи на слове {word}: {e}")
                return 0

    # Функция детекции китайских иероглифов
    def has_chinese(self, text):
        return any('\u4e00' <= char <= '\u9fff' for char in str(text))

    def has_unexpected_chars(self, text):
        # 1. Проверка на китайские иероглифы
        has_chinese = any('\u4e00' <= char <= '\u9fff' for char in str(text))
        if has_chinese:
            # Ищем, какой именно символ прикинулся иероглифом
            offender = [c for c in str(text) if '\u4e00' <= c <= '\u9fff']
            print(f"DEBUG: Нашел 'иероглиф' -> {offender}")
            return True

        # 2. Проверка на "мусорные" символы:
        # Ищем всё, что НЕ: латиница, кириллица, цифры, пробелы и пунктуация (. , ! ? - : ; ( ) [ ] " ') - это белый список
        # Регулярка [^...] означает "всё, кроме перечисленного"
     #   bad_chars = re.findall(r'[^a-zA-Zа-яА-ЯёЁ0-9\s\.,!\?\-\:\;\(\)\[\]"\'«»—\/\*\–\„\“]', str(text))
        bad_chars = re.findall(r"[^a-zA-Zа-яА-ЯёЁ0-9\s.,!?:;()\[\]\"'«»/\*„“\-\——–]", str(text))  # ^ - все, КРОМЕ
        # Исключаем из "плохих" символов перенос строки, чтобы не ломать парсинг
        bad_chars = [c for c in bad_chars if c not in ['\n', '\r']]
        if bad_chars:
            print(f"DEBUG: Нашел плохие символы -> {bad_chars}")

        return has_chinese or len(bad_chars) > 0

    def process_all_words(self, input_path, results_path, system_prompt, judge_system_prompt):
        """Читает список слов и сохраняет переводы по одному."""
        # Список для накопления строк таблицы
        self.rows = []
        # Список значимых колонок
        content_cols = ["Translation", "Examples", "Phrases"]

        with open(input_path, 'r', encoding='utf-8') as f:
            words_to_process = [line.strip() for line in f if line.strip()]

        print(f"Начинаю перевод {len(words_to_process)} слов в таблицу...")

        total_start_time = time.time()  # Общее время
        word_times = []  # Список для хранения времени каждого слова

        for word in words_to_process:
            word_start = time.time()
            try:
                max_attempts = 3
                attempt = 0
                success = False
                final_raw = ""
                final_score = 0

                # --- ЦИКЛ RETRY LOOP для рестарта перевода, если он неудовлетворителен ---
                while attempt < max_attempts and not success:
                    attempt += 1

                    # Первая попытка: железная строгость. Пересдача: добавляем "вариативность", чтобы выйти из тупика
                    curr_temp = 0.1 if attempt == 1 else 0.7

                    current_prompt = system_prompt
                    if attempt > 1:
                        # Усиливаем требования при повторной попытке
                        current_prompt = system_prompt + f"\n\nCRITICAL ERROR: Your previous translation for '{word}' was REJECTED as a hallucination. " \
                                                         f"DO NOT use non-existent words like 'себяг'. Use REAL Russian dictionary definitions." \
                                                         f"All [T], [E], [P] sections must be present. NO Chinese. Presence of the target word in examples."

                    # 1. Генерируем "сырой" ответ, с заменой температуры и сэмплами, если попыток > 1
                    final_raw = self.translate_word(word, current_prompt, temperature=curr_temp)

                    # 2. ТЕХНИЧЕСКАЯ ПРОВЕРКА НАЛИЧИЯ ТЕГОВ (tags_present)
                    tags_present = all(tag in final_raw for tag in ["[T]", "[E]", "[P]"])

                    # 3. ВЫЗЫВАЕМ СУДЬЮ
                    final_score = self.judge_adequacy(word, final_raw, judge_system_prompt)

                    # 4. Проверяем на неожиданные символы
                    unexpected_chars = self.has_unexpected_chars(final_raw)

                    # УСЛОВИЕ УСПЕХА: Score >= 4 + все теги на месте + нет китайского
                    fails = []
                    if not tags_present: fails.append("пустые теги")
                    if unexpected_chars: fails.append("неожиданные символы/иероглифы")
                    if final_score < 4: fails.append(f"низкий score {final_score}")

                    if not fails:  # Если список пуст — значит всё отлично
                        success = True
                        print(f"  [✓] {word}: Успех с попытки {attempt}")
                    else:
                        # Соединяем все ошибки через запятую
                        reason = ", ".join(fails)
                        print(f"  [!] {word}: Попытка {attempt} не удалась ({reason}).")

                # # # --- РАЗБИВКА ТЕКСТА ---

                # # ПАРСИНГ
                # Ищем блоки с помощью "заглядывания вперед" (?=\[|$) :
                # Бери весь текст, пока не встретишь ЛИБО начало следующего тега [, ЛИБО самый конец сообщения $
                # Это позволяет найти текст, даже если это последний блок в ответе
                t_match = re.search(r'\[T\]\s*(.*?)\s*(?=\[|$)', final_raw, re.DOTALL)
                e_match = re.search(r'\[E\]\s*(.*?)\s*(?=\[|$)', final_raw, re.DOTALL)
                p_match = re.search(r'\[P\]\s*(.*?)\s*(?=\[|$)', final_raw, re.DOTALL)

                # Безопасно извлекаем текст. Если блок найден (.group(1)), чистим его (.strip()), если нет — пишем ошибку.
                res_t = t_match.group(1).strip() if t_match else "Parsing error"
                res_e = e_match.group(1).strip() if e_match else "Parsing error"
                res_p = p_match.group(1).strip() if p_match else "Parsing error"

                # Определяем финальный скор для таблицы
                if not success:
                    status = "⚠ Галлюцинация" if final_score <= 2 else "⚠ Требует правки"
                else:
                    status = "✓ OK"

                # # Если судья поставил 1-2, помечаем прямо в ячейке перевода
                # display_translation = f"[{final_score}/5] {res_t}" if final_score <= 2 else res_t

                # Записываем в итоговый список (добавляем новые колонки)
                self.rows.append({
                    "English Word": word.upper(),
                    "Translation": res_t,
                    "Examples": res_e,
                    "Phrases": res_p,
                    "Score": final_score,  # колонка с цифрой
                    "Status": status  # колонка с вердиктом
                })

                # Сохраняем
                df = pd.DataFrame(self.rows)
                df.to_excel(results_path, index=False,
                            engine='openpyxl')  # сохраняем отформатированные данные с помощью openpyxl

                # --- блок для красивого форматирования ---
                wb = load_workbook(results_path)
                ws = wb.active
                # TODO Если заметишь замедление, просто вынеси блок с load_workbook и Alignment за пределы цикла for, чтобы отформатировать всё один раз в самом конце. Но пока слов немного — оставляй внутри для надежности.

                # Задаем ширину колонок: Слово(15), Перевод(25), Остальное(50)
                dims = {'A': 25, 'B': 35, 'C': 75, 'D': 75, 'E': 10, 'F': 20}
                for col, value in dims.items():
                    ws.column_dimensions[col].width = value

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

                wb.save(results_path)
                print(f"[✓] {word} добавлено в таблицу")

            except Exception as e:
                print(f"[!] Ошибка на слове {word}: {e}")

            time.sleep(0.5)
            word_end = time.time()
            duration = word_end - word_start
            word_duration = time.time() - word_start - 0.5
            word_times.append(word_duration)
            print(f"  ⏱ Время обработки '{word}': {word_duration:.2f} сек.")

        total_duration = time.time() - total_start_time

        # --- БЛОК МЕТРИК ---
        df = pd.DataFrame(self.rows)  # опционально. чтобы данные были доступны

        # Список колонок для проверки (чтобы Score и Status не портили статистику)
        content_cols = ["Translation", "Examples", "Phrases"]

        # Сбор статистики
        total_words = len(df)

        # Считаем ошибки парсинга только в контентных колонках
        parsing_errors_by_col = (df[content_cols] == "Parsing error").sum()
        total_parsing_errors = parsing_errors_by_col.sum()

        # Считаем пустые ячейки (используем .map вместо устаревшего .applymap)
        empty_cells = df[content_cols].map(lambda x: str(x).strip() == "").sum().sum()

        # Считаем иероглифы (хотя бы один в любой ячейке строки)
        chinese_count = df.apply(lambda row: row.map(self.has_chinese).any(), axis=1).sum()

        # --- ПЕЧАТЬ ОТЧЕТА ---
        if word_times:
            avg_time = sum(word_times) / len(word_times)
            print(f"\n⏱ ТАЙМИНГИ {self.model}:")
            print(f"  • Всего затрачено: {total_duration / 60:.1f} мин.")
            print(f"  • Среднее на 1 слово: {avg_time:.2f} сек.")
            print(f"  • Среднее на {len(words_to_process)} слов: {avg_time * len(words_to_process):.1f} сек.")

        print(f"\n" + "=" * 45)
        print(f"📊 ИТОГОВЫЙ ОТЧЕТ ПО КАЧЕСТВУ (Слов: {total_words})")
        print(f"⎯" * 45)

        print(f"❌ ТЕХНИЧЕСКИЙ БРАК:")
        for col in content_cols:
            print(f"  • {col}: {parsing_errors_by_col[col]} ошибок парсинга")
        print(f"  • Пустых ячеек (пропуски): {empty_cells}")

        if 'Score' in df.columns:
            hallucinations = (df['Score'] <= 2).sum()
            print(f"\n🧠 КАЧЕСТВО (LLM Judge):")
            print(f"  • Средняя адекватность: {df['Score'].mean():.2f} / 5.0")
            print(f"  • Галлюцинации (Score 1-2): {hallucinations}")

        print(f"\n🇨🇳 ЛОКАЛИЗАЦИЯ:")
        print(f"  • Слов с иероглифами: {chinese_count}")

        # Итоговый процент брака (только по контентным колонкам)
        total_content_cells = total_words * len(content_cols)
        if total_content_cells > 0:
            error_rate = ((total_parsing_errors + empty_cells) / total_content_cells) * 100
            print(f"⎯" * 45)
            print(f"📈 ОБЩИЙ ПРОЦЕНТ ТЕХНИЧЕСКОГО БРАКА: {error_rate:.2f}%")

        print("=" * 45)


# --- ЗАПУСК ---
if __name__ == "__main__":

    extractor = BatchEnglishExtractor()

    """ Распознаем слова"""
    # Проверка существования папки перед запуском
    # if not os.path.exists(full_set_path):
    #     try:
    #         os.makedirs(full_set_path)
    #         print(f"Папка '{full_set_path}' создана. Пожалуйста, положите туда картинки.")
    #     except OSError:
    #         print(f"Не удалось создать папку {full_set_path}. Проверьте пути.")
     # Формируем файл со словами
    # else:
    #     extractor.process_folder(full_set_path, OUTPUT_FILE_FULL)

    """ Подготовка таблицы уникальных слов """
    # extractor.analyze_frequency(OUTPUT_FILE_FULL, top_n=20)
    # extractor.remove_top_noise(OUTPUT_FILE_FULL, noise_count=7)
    # extractor.analyze_frequency(OUTPUT_FILE_FULL, top_n=10)
    extractor.remove_duplicates(OUTPUT_FILE_FULL)
    extractor.analyze_frequency(OUTPUT_FILE_FULL, top_n=10)

    """ Распознаем слова VtM"""
   # # Проверка существования папки перед запуском
   #  if not os.path.exists(set_path_VtM):
   #      try:
   #          os.makedirs(set_path_VtM)
   #          print(f"Папка '{set_path_VtM}' создана. Пожалуйста, положите туда картинки.")
   #      except OSError:
   #          print(f"Не удалось создать папку {set_path_VtM}. Проверьте пути.")
   #  # Формируем файл со словами
   #  else:
   #      extractor.process_folder(set_path_VtM, OUTPUT_FILE_VtM)
   #
   #  """ Подготовка таблицы уникальных слов """
   #  extractor.analyze_frequency(OUTPUT_FILE_VtM, top_n=20)
   #  extractor.remove_top_noise(OUTPUT_FILE_VtM, noise_count=7)
   #  extractor.analyze_frequency(OUTPUT_FILE_VtM, top_n=10)
   #  extractor.remove_duplicates(OUTPUT_FILE_VtM)
   #  extractor.analyze_frequency(OUTPUT_FILE_VtM, top_n=10)


    prompt_en_8_examples = """You are a translation engine.
    RULES:
    1. ONLY English and Russian languages used.
    2. NO Chinese/Asian characters or translations.
    3. If you see Chinese in your thoughts, DELETE it.
    4. DO NOT use rare or non-existent meanings.
    5. DO NOT make words up
    6. START each section strictly with the tags [T], [E], and [P].
    7. Each section tag must be filled. There must not be parsing errors
    8. Examples and phrases are well-composed and correctly translated to Russian
    9. [E] and [P]: Examples contain ONLY ENGLISH
    10. [E] and [P]: Example translations contain ONLY RUSSIAN
    11. [P]: If there are no well-known/set phrase with this word, it's allowed to leave "Not found"

    FORMAT:
    [T]
    (3-5 the most common Russian translations of this word, each on a new line)
    [E]
    (3 example phrases or full sentences using this word IN ENGLISH, each on a new line + hyphen + russian translation)
    [P]
    (1-3 set/well-known/slang phrases, or idioms (if they exist) using this word IN ENGLISH, each on a new line + hyphen + russian translation)

    EXAMPLE:
    Input: quizzically
    Output:
    [T]
    насмешливо
    чудаковато
    с недоумением
    [E]
    She was sitting with her head quizzically tilted - она сидела с головой, наклоненной чудаковатым образом
    [P]
    To look quizzically at someone - взглянуть на кого-то с недоумением"""

    judge_prompt_en_3 = """You are a ruthless Linguistic Auditor.
    Your goal is to find errors in a dictionary entry.

    CRITICAL ERRORS
    Score 1:
    - The Russian translation is a made-up word (hallucination like 'себяг').
    - The Russian translation is not found or empty
    - "Parsing error" in any section ([T], [E], or [P])

    Score 2:
    - Any Chinese/Asian characters found in any section ([T], [E], or [P]).
    - The Russian translation is completely unrelated to the English word or phrase.
    - Numbers found in any section
    - ([E] and [P] sections) the examples do not contain the word

    MINOR ERRORS (Score 3):
    - All 3 sections ([T], [E], or [P]) are present.
    - ([T] section) Only one translation provided instead of 3-5.
    - Russian grammar is broken.
    - The Russian translation is not the most common to the word
    - ([E] and [P] sections) Russian translation of phrases is incorrect
    - [P] section may be empty or contain "Not found"

    GOOD (Score 4):
    - All 3 sections ([T], [E], or [P]) are present and accurate.
    - ONLY English and Russian languages used.
    - ([T] section) at least 2 accurate Russian words, corresponding the most common translation of the word.
    - ([E] section) at least 3 suitable well-written examples provided in english with correct Russian translation
    - [P] section may be empty or contain "Not found"

    PERFECT (Score 5):
    - All 3 sections ([T], [E], or [P]) are present and accurate.
    - ONLY English and Russian languages used.
    - ([T] section) at least 3 accurate Russian words, corresponding the most common translation of the word.
    - ([E] section) at least 3 suitable well-written examples provided in english with correct Russian translation
    - ([P] section) at least 1 suitable phrase provided in english with correct Russian translation

    Respond ONLY with a single number from 1 to 5."""

    prompt_en_9 = """You are a translation engine.
    Your goal is to create a high-quality educational dictionary for advanced English learners. 
    Focus on natural, modern Russian and English and precise context
    RULES:
    1. ONLY English and Russian languages used.
    2. NO Chinese/Asian characters or translations.
    3. If you see Chinese in your thoughts, DELETE it.
    4. DO NOT use rare or non-exisent meanings.
    5. DO NOT make words up
    6. START each section strictly with the tags [T], [E], and [P].
    7. Each section tag must be filled. There must not be parsing errors
    8. Examples and phrases are well-composed and correctly translated to Russian
    9. [E] and [P]: Examples contain ONLY ENGLISH
    10. [E] and [P]: Example translations contain ONLY RUSSIAN
    11. [P]: Try your best to find at least one common collocation, idiom or set phrase. Only use 'Not found' if the word is extremely technical or has no established collocations.
    12. Context priority: In the [T] section, provide translations that cover different nuances of the word (e.g., both verb and noun forms if applicable).

    FORMAT:
    [T]
    (3-5 the most common Russian translations of this word, each on a new line)
    [E]
    (3 example phrases or full sentences using this word IN ENGLISH, each on a new line + hyphen + russian translation)
    [P]
    (1-3 set phrases, idioms, or collocations (if they exist) using this word IN ENGLISH, each on a new line + hyphen + russian translation)

    EXAMPLE:
    Input: fowl
    Output:
    [T]
    птица
    дичь
    пернатые
    [E]
    Wild fowl migrate south for the winter - Дикие птицы мигрируют на юг на зиму
    [P]
    Neither fish nor fowl - ни рыба ни мясо"""

    judge_prompt_en_4 = """You are a ruthless Linguistic Auditor.
    Your goal is to find errors in a dictionary entry.

    CRITICAL ERRORS
    Score 1:
    - The Russian translation is a made-up word (hallucination like 'себяг').
    - The Russian translation is not found or empty
    - "Parsing error" in any section ([T], [E], or [P])

    Score 2:
    - Any Chinese/Asian characters found in any section ([T], [E], or [P]).
    - The Russian translation is completely unrelated (hallucination).
    - Numbers found in any section
    - ([E] and [P] sections) the examples do not contain the word

    MINOR ERRORS (Score 3):
    - All 3 sections ([T], [E], or [P]) are present.
    - ([T] section) Only one translation provided instead of 3-5.
    - Russian grammar is broken.
    - The Russian translation is not the most common to the word
    - ([E] and [P] sections) Russian translation of phrases is incorrect

    GOOD (Score 4):
    - [T] and [E] sections are present and all requirements for them are met.
    - ONLY English and Russian languages used.
    - ([T] section) at least 2 accurate Russian words, corresponding the most common translation of the word.
    - ([E] section) at least 3 suitable well-written examples provided in english with correct Russian translation
    - [P] section contains accurate set phrases, idioms, or collocations  OR 'Not found' (if no common phrases exist). 
    - If [T] and [E] are excellent, 'Not found' in [P] is NOT a reason to lower the score below 4

    PERFECT (Score 5):
    - All 3 sections ([T], [E], or [P]) are present and all requirements for them are met.
    - ONLY English and Russian languages used.
    - ([T] section) at least 3 accurate Russian words, corresponding the most common translation of the word.
    - ([E] section) at least 3 suitable well-written examples provided in english with correct Russian translation
    - ([P] section) at least 1 suitable set phrase, idiom, or collocation is provided in english with correct Russian translation

    Respond ONLY with a single number from 1 to 5."""


    translator = EnglishContextTranslator(api_key=PT_API_KEY)

    # translator.process_all_words(OUTPUT_FILE_VtM, EXCEL_FILE_VtM, prompt_en_8_examples, judge_prompt_en_3)
    #translator.process_all_words(OUTPUT_FILE_VtM1, EXCEL_FILE_VtM1, prompt_en_9, judge_prompt_en_4)
    #translator.process_all_words(OUTPUT_FILE_VtM2, EXCEL_FILE_VtM2, prompt_en_9, judge_prompt_en_4)
    #translator.process_all_words(OUTPUT_FILE_VtM3, EXCEL_FILE_VtM3, prompt_en_9, judge_prompt_en_4)
    #translator.process_all_words(OUTPUT_FILE_VtM4, EXCEL_FILE_VtM4, prompt_en_9, judge_prompt_en_4)
    # translator.process_all_words(OUTPUT_FILE_VtM5, EXCEL_FILE_VtM5, prompt_en_9, judge_prompt_en_4)

   # translator.process_all_words(OUTPUT_FILE_TINY, EXCEL_FILE_TINY, prompt_en_9, judge_prompt_en_4)



    """
    Идея: Если хочешь идеала, можно добавить в current_prompt для 3-й попытки фразу: "If you are failing, try to look for a completely different meaning of the word."
    """


"""
1. Создайте отдельный класс (в том же файле или новом):
python
class GeminiTranslator:
    def __init__(self, api_key):
        self.client = genai.Client(api_key=api_key)
        self.model = "gemini-2.0-flash"

    def translate(self, word, system_prompt):
        try:
            response = self.client.models.generate_content(
                model=self.model,
                config=types.GenerateContentConfig(
                    system_instruction=system_prompt,
                    temperature=0.1
                ),
                contents=f"Translate: {word}",
            )
            return response.text
        except Exception as e:
            return f"API Error: {e}"
Use code with caution.

2. В основном "раздутом" классе просто добавьте его в __init__:
python
class BatchEnglishExtractor:
    def __init__(self, api_key, ...):
        # ... your old logic ...
        self.translator = GeminiTranslator(api_key) # Создаем помощника

    def process_all_words(self, ...):
        # Вызывайте через помощника:
        raw_response = self.translator.translate(word, self.system_prompt)
"""