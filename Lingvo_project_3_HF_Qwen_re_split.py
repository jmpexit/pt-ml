import os
import re
import torch
import easyocr  # <--- Замена Paddle
from spellchecker import SpellChecker
from tqdm.auto import tqdm
from openpyxl.styles import Alignment
from openpyxl import load_workbook

import transformers
from transformers import AutoTokenizer, AutoModelForCausalLM, BitsAndBytesConfig

from envs import OPENROUTER_TOKEN, HF_TOKEN

# --- ВАШИ ПУТИ (СОХРАНЕНЫ) ---
BASE_PATH = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo"
train_path = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo\train\train_set"
test_path = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo\test"
OUTPUT_FILE = os.path.join(BASE_PATH, "submission_final.csv")
RESULTS_FILE = os.path.join(BASE_PATH, "translated_csv_dictionary.csv")
EXCEL_FILE = os.path.join(BASE_PATH, "MyVocabulary.xlsx")


class BatchEnglishExtractor:
    def __init__(self):
        print("--- Инициализация ---")

        # 1. Используем PyTorch для проверки GPU (так как он у вас точно работает)
        self.use_gpu = torch.cuda.is_available()
        if self.use_gpu:
            print(f"[✓] GPU обнаружена: {torch.cuda.get_device_name(0)}")
        else:
            print("[x] GPU не найдена. Работаю на CPU.")

        # 2. Инициализация EasyOCR (вместо PaddleOCR)
        # verbose=False убирает лишний шум в консоли
        self.reader = easyocr.Reader(['en'], gpu=self.use_gpu, verbose=False)

        # 3. Инициализация словаря
        self.spell = SpellChecker()

    def _is_valid_english(self, text):
        """Фильтр 'Двойной замок': Unicode + Словарь"""
        clean_text = re.sub(r'[^\w\s]', '', text).strip()

        # Пропускаем мусор и короткие слова
        if not clean_text or clean_text.isdigit() or len(clean_text) < 2:
            return False

        # Шаг Б: Unicode-ловушка (Кириллица)
        if re.search(r'[\u0400-\u04ff]', text):
            return False

        # Шаг В: Словарная проверка
        words_to_check = [clean_text, clean_text.lower()]
        # known() возвращает список найденных слов. Если список не пуст -> True
        return bool(self.spell.known(words_to_check))

    def process_folder(self, folder_path, output_file):
        valid_extensions = ('.jpg', '.jpeg', '.png', '.bmp')

        if not os.path.exists(folder_path):
            print(f"ОШИБКА: Папка {folder_path} не найдена.")
            return

        files = [f for f in os.listdir(folder_path) if f.lower().endswith(valid_extensions)]
        print(f"Найдено изображений: {len(files)}")

        with open(output_file, "w", encoding="utf-8") as f_out:
            # Используем tqdm для красивой полоски прогресса
            for filename in tqdm(files, desc="Обработка"):
                img_path = os.path.join(folder_path, filename)

                try:
                    # EasyOCR возвращает список кортежей: (bbox, text, confidence)
                    result = self.reader.readtext(img_path)

                    if not result:
                        continue

                    # Сбор данных
                    page_words = []
                    for (bbox, text, prob) in result:
                        # EasyOCR bbox формат: [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
                        # Берем y1 (верхний левый угол) для сортировки
                        y_coord = bbox[0][1]
                        page_words.append({"text": text, "y": y_coord})

                    # Сортировка сверху вниз (как в вашем оригинале)
                    page_words.sort(key=lambda x: x["y"])

                    # Фильтрация и запись
                    for item in page_words:
                        word = item["text"]
                        if self._is_valid_english(word):
                            f_out.write(word + "\n")

                except Exception as e:
                    print(f"\nОшибка в файле {filename}: {e}")

        print(f"\nГотово! Все слова сохранены в: {output_file}")


"""Очистка и сбор статистики"""
from collections import Counter
import pandas as pd

def analyze_frequency(input_file, top_n=10):
    """Считает частотность и выводит статистику в консоль."""
    with open(input_file, 'r', encoding='utf-8') as file:
        words = [line.strip() for line in file if line.strip()]

    # Считаем частотность
    word_counts = Counter(words)

    # Сортируем по популярности (от самых частых к редким)
    common_words = word_counts.most_common()

    print(f"--- Статистика файла: {input_file} ---")
    print(f"Всего слов (с повторами): {len(words)}")
    print(f"Уникальных слов: {len(word_counts)}")
    print(f"\nТоп-{top_n} частых слов:")

    for word, count in common_words[:top_n]:
        print(f"{word}: {count}")

    return common_words


def remove_top_noise(input_file, noise_count=7):
    """Находит топ N самых частых слов и удаляет их из файла полностью."""
    with open(input_file, 'r', encoding='utf-8') as file:
        all_words = [line.strip() for line in file if line.strip()]

    # Находим, какие именно слова являются шумом
    word_counts = Counter(all_words)
    noise_words = set(word for word, count in word_counts.most_common(noise_count))
    # Когда делаем if word not in noise_words, Python в случае со списком каждый раз пробегает по нему.
    # В случае с set (множеством) поиск происходит мгновенно. На 600 словах разницы нет, но на 10 000+
    # это станет заметно.

    print(f"Удаляем шум: {', '.join(noise_words)}")

    # Оставляем только те слова, которых нет в списке шума
    filtered_words = [word for word in all_words if word not in noise_words]

    with open(input_file, 'w', encoding='utf-8') as f:
        for word in filtered_words:
            f.write(f"{word}\n")

    print(f"Файл {input_file} очищен от шума")
    return filtered_words # Опционально

def remove_duplicates(input_file):
    """Читает файл и возвращает список уникальных слов в порядке их появления."""
    with open(input_file, 'r', encoding='utf-8') as file:
        words = [line.strip() for line in file if line.strip()]

    # Использование dict.fromkeys сохраняет порядок первого появления слова
    unique_words = list(dict.fromkeys(words))

    with open(input_file, 'w', encoding='utf-8') as f:
        for word in unique_words:
            f.write(f"{word}\n")

    print(f"Файл {input_file} успешно обновлен.")

    return unique_words # Опционально


def translate_word(word):
    """Генерирует перевод и справку для одного слова."""
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"Переведи слово: {word}"}
    ]

    text = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
    inputs = tokenizer([text], return_tensors="pt").to(model.device)

    # Генерируем ответ
    outputs = model.generate(**inputs, max_new_tokens=800, temperature=0.7)

    # Декодируем срезаем промпт, оставляем только ответ)
    response = tokenizer.decode(outputs[0][len(inputs.input_ids[0]):], skip_special_tokens=True)
    return response

def process_all_words(input_path, results_path):
    """Читает список слов и сохраняет переводы по одному."""

    # Список для накопления строк таблицы
    rows = []

    with open(input_path, 'r', encoding='utf-8') as f:
        words_to_process = [line.strip() for line in f if line.strip()]

    print(f"Начинаю перевод {len(words_to_process)} слов в таблицу...")

    for word in words_to_process:
        try:
            raw_response = translate_word(word) # функция перевода

            # --- РАЗБИВКА ТЕКСТА НА КОЛОНКИ ---
            # # Парсинг
            parts = re.split(r'#### \d+\.\s?.*?\n', raw_response) # Регулярка ищет "#### номер. Заголовок" и убирает это
            parts = [p.strip() for p in parts if p.strip()]  # Чистим пустые элементы, которые оставляет split

            translation = parts[0] if len(parts) > 0 else "Ошибка парсинга"
            examples = parts[1] if len(parts) > 1 else "Ошибка парсинга"
            idioms = parts[2] if len(parts) > 2 else "Ошибка парсинга"

            # Добавляем строку в список
            rows.append({
                "English Word": word.upper(),
                "Translation": translation,
                "Examples ": examples,
                "Phrases": idioms
            })

            # Сохраняем
            df = pd.DataFrame(rows)
          #  df.to_excel(results_path, index=False)
            df.to_excel(results_path, index=False, engine='openpyxl') #  сохраняем отформатированные данные с помощью openpyxl

            # --- блок для красивого форматирования ---
            wb = load_workbook(results_path)
            ws = wb.active

            # Задаем ширину колонок: Слово(15), Перевод(25), Остальное(50)
            dims = {'A': 15, 'B': 25, 'C': 60, 'D': 60}
            for col, value in dims.items():
                ws.column_dimensions[col].width = value

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

            wb.save(results_path)
            print(f"[✓] {word} успешно добавлено в таблицу")

        except Exception as e:
            print(f"[!] Ошибка на слове {word}: {e}")

#    df = pd.DataFrame(rows)

    # --- БЛОК МЕТРИК ---

    def has_chinese(text):
        # Диапазон иероглифов в Unicode
        return any('\u4e00' <= char <= '\u9fff' for char in str(text))

    # Считаем количество строк, где проскочил китайский
    chinese_errors = df.apply(lambda row: row.map(has_chinese).any(), axis=1).sum()

    # Считаем, сколько раз встретилась фраза "Ошибка парсинга" в каждой колонке
    error_summary = (df == "Ошибка парсинга").sum()
    total_errors = error_summary.drop("English Word", errors='ignore').sum()
    total_words = len(df)

    print(f"\n" + "=" * 40)
    print(f"📊 ОТЧЕТ ПО КАЧЕСТВУ (Всего слов: {total_words})")
    print(f"⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯")
    # Выводим статистику по иероглифам
    print(f"• Иероглифы найдены в {chinese_errors} словах")
    print(f"⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯")
    # Выводим статистику по ошибке парсинга каждой колонке
    for column, count in error_summary.items():
        if column != "English Word":  # Слово у нас всегда есть
            print(f"• {column}: {count} ошибок")
    # Считаем общий процент
    error_rate = (total_errors / (total_words * 3)) * 100  # 3 колонки с контентом
    print(f"⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯")
    print(f"📈 Общий процент ошибок: {error_rate:.2f}%")
    print("=" * 40)

    print(f"\n--- Завершено. Таблица: {results_path} ---")


    """ WRAP
    from openpyxl.styles import Alignment

# ... (после сохранения df.to_excel) ...
with pd.ExcelWriter(output_excel, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    sheet = writer.sheets['Sheet1']
    for row in sheet.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    """


# --- ЗАПУСК ---
if __name__ == "__main__":
    # """ Распознаем слова """
    # # Проверка существования папки перед запуском
    # if not os.path.exists(test_path):
    #     try:
    #         os.makedirs(test_path)
    #         print(f"Папка '{test_path}' создана. Пожалуйста, положите туда картинки.")
    #     except OSError:
    #         print(f"Не удалось создать папку {test_path}. Проверьте пути.")
    # # Формируем файл со словами
    # else:
    #     extractor = BatchEnglishExtractor()
    #     extractor.process_folder(test_path, OUTPUT_FILE)
    #
    # """ Подготовка таблицы уникальных слов """
    # remove_top_noise(OUTPUT_FILE, noise_count=7)
    # remove_duplicates(OUTPUT_FILE)
    # analyze_frequency(OUTPUT_FILE, top_n=20)

    """ Подключаем GPT 
    
    pip install huggingface_hub
    pip install bitsandbytes accelerate
    """

    print(f"CUDA Available: {torch.cuda.is_available()}")
    print(f"Current Device: {torch.cuda.get_device_name(0)}")
    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"[✓] Using device: {device}")

    #model_name = "meta-llama/Meta-Llama-3-8B-Instruct"  # требует верификафии перс.данных
   # model_name = "mistralai/Mistral-7B-Instruct-v0.3"  # прямой конкурент Llama. быстрая и хорошо понимает инструкции - не идет загрузка
    model_name = "Qwen/Qwen2.5-7B-Instruct"  # Очень точные переводы и примеры
   # model_name = "Vikhert/Vikhr-Llama-3.1-8B-Instruct-r24"  # "дообученная" Llama, которую российские разработчики адаптировали под русский язык

    # Создаем конфигурацию для 8-битного режима
    bnb_config = BitsAndBytesConfig(
        load_in_8bit=True
    )

    tokenizer = AutoTokenizer.from_pretrained(model_name, token=HF_TOKEN) # TODO объяснить
    model = AutoModelForCausalLM.from_pretrained(
        model_name,
      #  torch_dtype=torch.float16,
        device_map="auto",
        quantization_config=bnb_config,  # если не хватает памяти. или load_in_4bit=True
        low_cpu_mem_usage=True, # Загрузка по частям и  создание пустого каркаса
        token=HF_TOKEN
    )


    # system_prompt = """You are a professional linguist and historian. # TODO куда подсунуть?
    # When translating a word:
    # 1. Provide a precise Russian translation (5 examples maximum).
    # 2. Give 3 vivid usage examples in context.
    # 3. Briefly explain the origin (etymology) or a fun historical fact about the word.
    # Keep the tone helpful and academic yet conversational."""

    system_prompt = """Ты — профессиональный лингвист. Переводи английские слова для личного словаря. Начинай содержимое каждого пункта с новой строки сразу после заголовка. Пиши только на русском и английском. Избегай иероглифов.
    Для каждого слова строго соблюдай структуру:
    #### 1. Перевод на русский: только русские слова через запятую. постарайся найти 3-5 наиболее используемых вариантов.
    #### 2. Примеры: 3 предложения на АНГЛИЙСКОМ, каждое - на новой строке, без перевода.
    #### 3. Устоявшиеся выражения: если слово входит в какое-либо устоявшееся выражение, приведи пример на английском, с интерпретацией на русском и исторической справкой на русском."""
#    ВАЖНО: Пиши только на русском и английском. Избегай иероглифов и лишних пояснений. Отвечай структурировано и интересно."""


    process_all_words(OUTPUT_FILE, EXCEL_FILE)

#     # yield для файла # TODO
#
#     def word_generator(input_path):
#     """Генератор, который читает файл и выдает по одному слову."""
#     with open(input_path, 'r', encoding='utf-8') as f:
#         for line in f:
#             word = line.strip()
#             if word:
#                 yield word
#
# def translation_stream(input_path):
#     """Генератор, который получает слово и возвращает готовый перевод."""
#     for word in word_generator(input_path):
#         print(f"--- Процессинг: {word} ---")
#         try:
#             # Твоя тяжелая функция с моделью
#             result = translate_word(word)
#             yield word, result
#         except Exception as e:
#             print(f"Ошибка на {word}: {e}")
#             continue
#
# # --- ОСНОВНОЙ ЗАПУСК ---
# if __name__ == "__main__":
#     # Теперь мы просто 'подписываемся' на поток переводов
#     for word, translation in translation_stream(OUTPUT_FILE):
#         with open(RESULTS_FILE, 'a', encoding='utf-8') as out_f:
#             out_f.write(f"\n{'='*30}\nWORD: {word.upper()}\n{translation}\n{'='*30}\n")
#         print(f"[✓] {word} записано.")




#   # Читаем слова по N штук
#   with open(OUTPUT_FILE, 'r', encoding='utf-8') as fIle:
#       words_to_process = [line.strip() for line in fIle if line.strip()][:5]
#
# #  words_list = "\n".join([f"- {w}" for w in words])
  #
  #   for word in words_to_process:
  #       messages = [
  #           {"role": "system", "content": system_prompt},
  #           {"role": "user", "content": f"Переведи слово: {word}"}
  #       ]
  #
  #       text = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
  #       model_inputs = tokenizer([text], return_tensors="pt").to(model.device)
  #
  #       generated_ids = model.generate(**model_inputs, max_new_tokens=512, temperature=0.7)
  #       response = tokenizer.batch_decode(generated_ids[:, model_inputs.input_ids.shape[1]:], skip_special_tokens=True)[0]
  #
  #       print(f"\n=== {word.upper()} ===\n{response}")


    # """ Llama-3 Prompt """
    # messages = [
    #     {"role": "system",
    #      "content": "You are a professional linguist. Translate English words to Russian and provide 3 usage examples in Russian for each."},
    #     {"role": "user", "content": f"Translate these words:\n{words_list}"},
    # ]
    # prompt = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)

    # # Generation
    # inputs = tokenizer(prompt, return_tensors="pt").to(model.device)
    #
    # # Increase max_new_tokens for 10 words + examples
    # outputs = model.generate(
    #     **inputs,
    #     max_new_tokens=1500,
    #     temperature=0.6,
    #     top_p=0.9,
    #     eos_token_id=tokenizer.eos_token_id
    # )
    #
    # print(tokenizer.decode(outputs[0][len(inputs["input_ids"][0]):], skip_special_tokens=True))