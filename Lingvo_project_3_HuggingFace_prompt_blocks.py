import os
import re
import torch
import easyocr  # <--- Замена Paddle
from spellchecker import SpellChecker
from tqdm.auto import tqdm
from openpyxl.styles import Alignment
from openpyxl import load_workbook

import transformers
from transformers import AutoTokenizer, AutoModelForCausalLM, BitsAndBytesConfig

from envs import OPENROUTER_TOKEN, HF_TOKEN

# --- ВАШИ ПУТИ (СОХРАНЕНЫ) ---
BASE_PATH = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo"
train_path = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo\train\train_set"
test_path = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo\test"
OUTPUT_FILE = os.path.join(BASE_PATH, "submission_final.csv")
RESULTS_FILE = os.path.join(BASE_PATH, "translated_csv_dictionary.csv")
EXCEL_FILE = os.path.join(BASE_PATH, "MyVocabulary.xlsx")


class BatchEnglishExtractor:
    def __init__(self):
        print("--- Инициализация ---")

        # 1. Используем PyTorch для проверки GPU (так как он у вас точно работает)
        self.use_gpu = torch.cuda.is_available()
        if self.use_gpu:
            print(f"[✓] GPU обнаружена: {torch.cuda.get_device_name(0)}")
        else:
            print("[x] GPU не найдена. Работаю на CPU.")

        # 2. Инициализация EasyOCR (вместо PaddleOCR)
        # verbose=False убирает лишний шум в консоли
        self.reader = easyocr.Reader(['en'], gpu=self.use_gpu, verbose=False)

        # 3. Инициализация словаря
        self.spell = SpellChecker()

    def _is_valid_english(self, text):
        """Фильтр 'Двойной замок': Unicode + Словарь"""
        clean_text = re.sub(r'[^\w\s]', '', text).strip()

        # Пропускаем мусор и короткие слова
        if not clean_text or clean_text.isdigit() or len(clean_text) < 2:
            return False

        # Шаг Б: Unicode-ловушка (Кириллица)
        if re.search(r'[\u0400-\u04ff]', text):
            return False

        # Шаг В: Словарная проверка
        words_to_check = [clean_text, clean_text.lower()]
        # known() возвращает список найденных слов. Если список не пуст -> True
        return bool(self.spell.known(words_to_check))

    def process_folder(self, folder_path, output_file):
        valid_extensions = ('.jpg', '.jpeg', '.png', '.bmp')

        if not os.path.exists(folder_path):
            print(f"ОШИБКА: Папка {folder_path} не найдена.")
            return

        files = [f for f in os.listdir(folder_path) if f.lower().endswith(valid_extensions)]
        print(f"Найдено изображений: {len(files)}")

        with open(output_file, "w", encoding="utf-8") as f_out:
            # Используем tqdm для красивой полоски прогресса
            for filename in tqdm(files, desc="Обработка"):
                img_path = os.path.join(folder_path, filename)

                try:
                    # EasyOCR возвращает список кортежей: (bbox, text, confidence)
                    result = self.reader.readtext(img_path)

                    if not result:
                        continue

                    # Сбор данных
                    page_words = []
                    for (bbox, text, prob) in result:
                        # EasyOCR bbox формат: [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
                        # Берем y1 (верхний левый угол) для сортировки
                        y_coord = bbox[0][1]
                        page_words.append({"text": text, "y": y_coord})

                    # Сортировка сверху вниз (как в вашем оригинале)
                    page_words.sort(key=lambda x: x["y"])

                    # Фильтрация и запись
                    for item in page_words:
                        word = item["text"]
                        if self._is_valid_english(word):
                            f_out.write(word + "\n")

                except Exception as e:
                    print(f"\nОшибка в файле {filename}: {e}")

        print(f"\nГотово! Все слова сохранены в: {output_file}")


"""Очистка и сбор статистики"""
from collections import Counter
import pandas as pd

def analyze_frequency(input_file, top_n=10):
    """Считает частотность и выводит статистику в консоль."""
    with open(input_file, 'r', encoding='utf-8') as file:
        words = [line.strip() for line in file if line.strip()]

    # Считаем частотность
    word_counts = Counter(words)

    # Сортируем по популярности (от самых частых к редким)
    common_words = word_counts.most_common()

    print(f"--- Статистика файла: {input_file} ---")
    print(f"Всего слов (с повторами): {len(words)}")
    print(f"Уникальных слов: {len(word_counts)}")
    print(f"\nТоп-{top_n} частых слов:")

    for word, count in common_words[:top_n]:
        print(f"{word}: {count}")

    return common_words


def remove_top_noise(input_file, noise_count=7):
    """Находит топ N самых частых слов и удаляет их из файла полностью."""
    with open(input_file, 'r', encoding='utf-8') as file:
        all_words = [line.strip() for line in file if line.strip()]

    # Находим, какие именно слова являются шумом
    word_counts = Counter(all_words)
    noise_words = set(word for word, count in word_counts.most_common(noise_count))
    # Когда делаем if word not in noise_words, Python в случае со списком каждый раз пробегает по нему.
    # В случае с set (множеством) поиск происходит мгновенно. На 600 словах разницы нет, но на 10 000+
    # это станет заметно.

    print(f"Удаляем шум: {', '.join(noise_words)}")

    # Оставляем только те слова, которых нет в списке шума
    filtered_words = [word for word in all_words if word not in noise_words]

    with open(input_file, 'w', encoding='utf-8') as f:
        for word in filtered_words:
            f.write(f"{word}\n")

    print(f"Файл {input_file} очищен от шума")
    return filtered_words # Опционально

def remove_duplicates(input_file):
    """Читает файл и возвращает список уникальных слов в порядке их появления."""
    with open(input_file, 'r', encoding='utf-8') as file:
        words = [line.strip() for line in file if line.strip()]

    # Использование dict.fromkeys сохраняет порядок первого появления слова
    unique_words = list(dict.fromkeys(words))

    with open(input_file, 'w', encoding='utf-8') as f:
        for word in unique_words:
            f.write(f"{word}\n")

    print(f"Файл {input_file} успешно обновлен.")

    return unique_words # Опционально


def translate_word(word, system_prompt):
    """Генерирует перевод и справку для одного слова."""
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"Переведи слово: {word}"}
    ]

    text = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
    inputs = tokenizer([text], return_tensors="pt").to(model.device)

    # Генерируем ответ
    outputs = model.generate(**inputs, max_new_tokens=800, temperature=0.1, do_sample=False) # низкая temperature=0.1. Для таблиц нам нужна максимальная строгость, а не творчество

    # Декодируем срезаем промпт, оставляем только ответ)
    response = tokenizer.decode(outputs[0][len(inputs.input_ids[0]):], skip_special_tokens=True)
    return response

def judge_adequacy(word, translation, judge_prompt):
    """Модель оценивает свою (или чужую) работу."""
    prompt = f"Word: {word}\nTranslation: {translation}\nRate adequacy (1-5):"

    messages = [
        {"role": "system", "content": judge_prompt},
        {"role": "user", "content": prompt}
    ]

    text = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
    inputs = tokenizer(text, return_tensors="pt").to(model.device)

    # Генерируем 2 токена (нам нужна только цифра)
    outputs = model.generate(**inputs, max_new_tokens=2, temperature=0.1, do_sample=False)

    score_raw = tokenizer.decode(outputs[0][len(inputs.input_ids[0]):], skip_special_tokens=True)

    match = re.search(r'\d', score_raw)
    return int(match.group()) if match else 0  # Возвращаем цифру или 0, если модель промолчала

def process_all_words(input_path, results_path, system_prompt):
    """Читает список слов и сохраняет переводы по одному."""

    # Список для накопления строк таблицы
    rows = []

    with open(input_path, 'r', encoding='utf-8') as f:
        words_to_process = [line.strip() for line in f if line.strip()]

    print(f"Начинаю перевод {len(words_to_process)} слов в таблицу...")

    for word in words_to_process:
        try:
            raw_response = translate_word(word, system_prompt) # функция перевода

            # # # --- РАЗБИВКА ТЕКСТА ---

            # # ПАРСИНГ
            # 1. Ищем блоки с помощью "заглядывания вперед" (?=\[|$) :
            # Бери весь текст, пока не встретишь ЛИБО начало следующего тега [, ЛИБО самый конец сообщения $
            # Это позволяет найти текст, даже если это последний блок в ответе
            t_match = re.search(r'\[T\]\s*(.*?)\s*(?=\[|$)', raw_response, re.DOTALL)
            e_match = re.search(r'\[E\]\s*(.*?)\s*(?=\[|$)', raw_response, re.DOTALL)
            p_match = re.search(r'\[P\]\s*(.*?)\s*(?=\[|$)', raw_response, re.DOTALL)

            # 2. Безопасно извлекаем текст. Если блок найден (.group(1)), чистим его (.strip()), если нет — пишем ошибку.
            res_t = t_match.group(1).strip() if t_match else "Ошибка парсинга"
            res_e = e_match.group(1).strip() if e_match else "Ошибка парсинга"
            res_p = p_match.group(1).strip() if p_match else "Ошибка парсинга"

            # 3. Добавляем результат в список строк для Excel
            rows.append({
                "English Word": word.upper(),
                "Translation": res_t,
                "Examples": res_e,
                "Phrases": res_p
            })

            # Сохраняем
            df = pd.DataFrame(rows)
          #  df.to_excel(results_path, index=False)
            df.to_excel(results_path, index=False, engine='openpyxl') #  сохраняем отформатированные данные с помощью openpyxl

            # --- блок для красивого форматирования ---
            wb = load_workbook(results_path)
            ws = wb.active
             # TODO Если заметишь замедление, просто вынеси блок с load_workbook и Alignment за пределы цикла for, чтобы отформатировать всё один раз в самом конце. Но пока слов немного — оставляй внутри для надежности.

            # Задаем ширину колонок: Слово(15), Перевод(25), Остальное(50)
            dims = {'A': 25, 'B': 40, 'C': 80, 'D': 80}
            for col, value in dims.items():
                ws.column_dimensions[col].width = value

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

            wb.save(results_path)
            print(f"[✓] {word} успешно добавлено в таблицу")

        except Exception as e:
            print(f"[!] Ошибка на слове {word}: {e}")

    # --- БЛОК МЕТРИК ---
  #  df = pd.DataFrame(rows)
    # Сбор статистики
    total_words = len(df)

    # Функция детекции китайских иероглифов
    def has_chinese(text):
        return any('\u4e00' <= char <= '\u9fff' for char in str(text))

    # Считаем иероглифы (хотя бы один в любой ячейке строки)
    chinese_words_count = df.apply(lambda row: row.map(has_chinese).any(), axis=1).sum()

    # Считаем ошибки парсинга по каждой колонке (кроме основного слова)
    parsing_errors_by_col = (df == "Ошибка парсинга").sum().drop("English Word", errors='ignore')
    total_parsing_errors = parsing_errors_by_col.sum()

    # Считаем пустые ячейки (где только пробелы или пусто)
    empty_cells = df.drop("English Word", axis=1, errors='ignore').applymap(lambda x: str(x).strip() == "").sum().sum()

    # # Считаем галлюцинации (оценка Судьи 1 или 2)
    # hallucinations = (df['Score'] <= 2).sum() if 'Score' in df.columns else 0

    # --- ПЕЧАТЬ ОТЧЕТА ---
    print(f"\n" + "=" * 45)
    print(f"📊 ИТОГОВЫЙ ОТЧЕТ ПО КАЧЕСТВУ (Слов: {total_words})")
    print(f"⎯" * 45)

    print(f"❌ ТЕХНИЧЕСКИЙ БРАК:")
    for column, count in parsing_errors_by_col.items():
        print(f"  • {column}: {count} ошибок парсинга")
    print(f"  • Пустых ячеек найдено: {empty_cells}")

    # print(f"\n🧠 КАЧЕСТВО КОНТЕНТА (LLM Judge):")
    # if 'Score' in df.columns:
    #     avg_score = df['Score'].mean()
    #     print(f"  • Средняя адекватность: {avg_score:.2f} / 5.0")
    #     print(f"  • Галлюцинации (Score 1-2): {hallucinations}")

    print(f"\n🇨🇳 ЛОКАЛИЗАЦИЯ:")
    print(f"  • Слов с иероглифами: {chinese_words_count}")

    # Итоговый Error Rate (Парсинг + Пустоты)
    # Делим на (кол-во слов * 3 колонки контента)
    total_content_cells = total_words * 3
    if total_content_cells > 0:
        error_rate = ((total_parsing_errors + empty_cells) / total_content_cells) * 100
        print(f"⎯" * 45)
        print(f"📈 ОБЩИЙ ПРОЦЕНТ ТЕХНИЧЕСКОГО БРАКА: {error_rate:.2f}%")

    print("=" * 45)
    print(f"Файл сохранен: {results_path}")



    print(f"\n--- Завершено. Таблица: {results_path} ---")


# --- ЗАПУСК ---
if __name__ == "__main__":
    # """ Распознаем слова """
    # # Проверка существования папки перед запуском
    # if not os.path.exists(test_path):
    #     try:
    #         os.makedirs(test_path)
    #         print(f"Папка '{test_path}' создана. Пожалуйста, положите туда картинки.")
    #     except OSError:
    #         print(f"Не удалось создать папку {test_path}. Проверьте пути.")
    # # Формируем файл со словами
    # else:
    #     extractor = BatchEnglishExtractor()
    #     extractor.process_folder(test_path, OUTPUT_FILE)
    #
    # """ Подготовка таблицы уникальных слов """
    # remove_top_noise(OUTPUT_FILE, noise_count=7)
    # remove_duplicates(OUTPUT_FILE)
    # analyze_frequency(OUTPUT_FILE, top_n=20)

    """ Подключаем GPT 
    pip install huggingface_hub
    pip install bitsandbytes accelerate
    """

    print(f"CUDA Available: {torch.cuda.is_available()}")
    print(f"Current Device: {torch.cuda.get_device_name(0)}")
    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"[✓] Using device: {device}")

    model_name = "Qwen/Qwen2.5-7B-Instruct"  # Очень точные переводы и примеры

    # Создаем конфигурацию для 8-битного режима
    bnb_config = BitsAndBytesConfig(
        load_in_8bit=True
    )

    tokenizer = AutoTokenizer.from_pretrained(model_name, token=HF_TOKEN) # TODO объяснить
    model = AutoModelForCausalLM.from_pretrained(
        model_name,
      #  torch_dtype=torch.float16,
        device_map="auto",
        quantization_config=bnb_config,  # если не хватает памяти. или load_in_4bit=True
        low_cpu_mem_usage=True, # Загрузка по частям и  создание пустого каркаса
        token=HF_TOKEN
    )

    # system_prompt = """You are a professional linguist. Use ONLY English and Russian. NO Chinese.
    # For each word, follow this EXACT structure:
    #
    # [TRANSLATION]
    # (Russian translation only, 3-5 variants)
    #
    # [EXAMPLES]
    # (3 English sentences only, no translation)
    #
    # [PHRASES]
    # (English idioms with Russian comments)
    # """

    prompt_en_1 = """You are a precise English-Russian dictionary bot. 
    Use ONLY English and Russian. NO Chinese characters.
    Structure your response EXACTLY as follows:

    [T]
    (3-5 Russian translations, comma separated)

    [E]
    (3 English example sentences, each on a new line, with russian translation, hyphen separated)

    [P]
    (3 English phrases or idioms with brief Russian comments)"""

    prompt_en_2 = """You are a precise English-Russian dictionary bot. 
    STRICT RULES:
    1. Use ONLY English and Russian. NO Chinese characters.
    2. DO NOT use rare or non-eexisent meanings.
    3. START each section strictly with the tags [T], [E], and [P].

    [T]
    (3-5 Russian translations, comma separated)

    [E]
    (3 English example sentences, each on a new line, with russian translation, hyphen separated)

    [P]
    (3 English phrases/idioms with Russian comments)"""

    prompt_en_3 = """You are a professional linguist. 
    STRICT RULES:
    1. Use ONLY English and Russian. NO Chinese characters or translations.
    2. DO NOT use rare or non-exisent meanings.
    3. START each section strictly with the tags [T], [E], and [P].

    [T]
    (3-5 Russian translations of this word, each translation on a new line)

    [E]
    (3 English example sentences using this word, each on a new line, with russian translation after hyphen)

    [P]
    (3 English well-known phrases or idioms using this word, each on a new line, with russian translation after hyphen)"""

    prompt_en_4 = """You are a translation engine. 
    RULES:
    1. ONLY English and Cyrillic Russian. 
    2. NO Chinese/Asian characters.
    3. If you see Chinese in your thoughts, DELETE it.

    FORMAT:
    [T]
    (Russian words only)
    [E]
    (English sentence - Russian translation)
    [P]
    (English idiom - Russian translation)"""

    prompt_en_5 = """You are a translation engine. 
    RULES:
    1. ONLY English and Cyrillic Russian. 
    2. NO Chinese/Asian characters.
    3. If you see Chinese in your thoughts, DELETE it.
    4. DO NOT use rare or non-exisent meanings.
    5. START each section strictly with the tags [T], [E], and [P].

    FORMAT:
    [T]
    (3-5 Russian translations of this word, each translation on a new line)
    [E]
    (3 English example sentences using this word, each on a new line, with russian translation after hyphen)
    [P]
    (3 English well-known phrases or idioms using this word, each on a new line, with russian translation after hyphen)"""



    prompt_ru_1 = """Ты — профессиональный лингвист. Переводи английские слова для личного словаря.
    Используй ТОЛЬКО русский и английский языки. КАТЕГОРИЧЕСКИ запрещено использовать иероглифы.
    Для каждого слова СТРОГО соблюдай структуру ответа:

    [T]
    (здесь только перевод на РУССКИЙ через запятую, 3-5 слов, в зависимости от количества разных значений слова)

    [E]
    (здесь 3 примера на английском, с переводом на русский через дефис)

    [P]
    (здесь устоявшиеся выражения с этим словом с комментариями на русском)
    """

    judge_prompt = """You are a strict linguistic auditor. 
    Evaluate the translation quality between English and Russian.
    Rate the 'Adequacy' from 1 to 5:
    5 - Perfect translation.
    4 - Good, but slightly rare meaning.
    3 - Correct word, but weird context.
    2 - Wrong meaning or invented word (hallucination).
    1 - Total nonsense or gibberish.

    Respond ONLY with a single number."""

    process_all_words(OUTPUT_FILE, EXCEL_FILE, prompt_en_4)
