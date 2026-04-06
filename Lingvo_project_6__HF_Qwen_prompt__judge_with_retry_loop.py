import os
import re
import torch
import easyocr  # <--- Замена Paddle
from spellchecker import SpellChecker
from tqdm.auto import tqdm
from openpyxl.styles import Alignment
from openpyxl import load_workbook

import transformers
from transformers import AutoTokenizer, AutoModelForCausalLM, BitsAndBytesConfig

from envs import OPENROUTER_TOKEN, HF_TOKEN

# --- ВАШИ ПУТИ (СОХРАНЕНЫ) ---
BASE_PATH = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo"
train_path = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo\train\train_set"
test_path = r"C:\Users\Julie\PycharmProjects\pt-ml\datasets\lingvo\test"
OUTPUT_FILE = os.path.join(BASE_PATH, "submission_final.csv")
RESULTS_FILE = os.path.join(BASE_PATH, "translated_csv_dictionary.csv")
EXCEL_FILE = os.path.join(BASE_PATH, "MyVocabulary.xlsx")


class BatchEnglishExtractor:
    def __init__(self):
        print("--- Инициализация ---")

        # 1. Используем PyTorch для проверки GPU (так как он у вас точно работает)
        self.use_gpu = torch.cuda.is_available()
        if self.use_gpu:
            print(f"[✓] GPU обнаружена: {torch.cuda.get_device_name(0)}")
        else:
            print("[x] GPU не найдена. Работаю на CPU.")

        # 2. Инициализация EasyOCR (вместо PaddleOCR)
        # verbose=False убирает лишний шум в консоли
        self.reader = easyocr.Reader(['en'], gpu=self.use_gpu, verbose=False)

        # 3. Инициализация словаря
        self.spell = SpellChecker()

    def _is_valid_english(self, text):
        """Фильтр 'Двойной замок': Unicode + Словарь"""
        clean_text = re.sub(r'[^\w\s]', '', text).strip()

        # Пропускаем мусор и короткие слова
        if not clean_text or clean_text.isdigit() or len(clean_text) < 2:
            return False

        # Шаг Б: Unicode-ловушка (Кириллица)
        if re.search(r'[\u0400-\u04ff]', text):
            return False

        # Шаг В: Словарная проверка
        words_to_check = [clean_text, clean_text.lower()]
        # known() возвращает список найденных слов. Если список не пуст -> True
        return bool(self.spell.known(words_to_check))

    def process_folder(self, folder_path, output_file):
        valid_extensions = ('.jpg', '.jpeg', '.png', '.bmp')

        if not os.path.exists(folder_path):
            print(f"ОШИБКА: Папка {folder_path} не найдена.")
            return

        files = [f for f in os.listdir(folder_path) if f.lower().endswith(valid_extensions)]
        print(f"Найдено изображений: {len(files)}")

        with open(output_file, "w", encoding="utf-8") as f_out:
            # Используем tqdm для красивой полоски прогресса
            for filename in tqdm(files, desc="Обработка"):
                img_path = os.path.join(folder_path, filename)

                try:
                    # EasyOCR возвращает список кортежей: (bbox, text, confidence)
                    result = self.reader.readtext(img_path)

                    if not result:
                        continue

                    # Сбор данных
                    page_words = []
                    for (bbox, text, prob) in result:
                        # EasyOCR bbox формат: [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
                        # Берем y1 (верхний левый угол) для сортировки
                        y_coord = bbox[0][1]
                        page_words.append({"text": text, "y": y_coord})

                    # Сортировка сверху вниз (как в вашем оригинале)
                    page_words.sort(key=lambda x: x["y"])

                    # Фильтрация и запись
                    for item in page_words:
                        word = item["text"]
                        if self._is_valid_english(word):
                            f_out.write(word + "\n")

                except Exception as e:
                    print(f"\nОшибка в файле {filename}: {e}")

        print(f"\nГотово! Все слова сохранены в: {output_file}")


"""Очистка и сбор статистики"""
from collections import Counter
import pandas as pd


def analyze_frequency(input_file, top_n=10):
    """Считает частотность и выводит статистику в консоль."""
    with open(input_file, 'r', encoding='utf-8') as file:
        words = [line.strip() for line in file if line.strip()]

    # Считаем частотность
    word_counts = Counter(words)

    # Сортируем по популярности (от самых частых к редким)
    common_words = word_counts.most_common()

    print(f"--- Статистика файла: {input_file} ---")
    print(f"Всего слов (с повторами): {len(words)}")
    print(f"Уникальных слов: {len(word_counts)}")
    print(f"\nТоп-{top_n} частых слов:")

    for word, count in common_words[:top_n]:
        print(f"{word}: {count}")

    return common_words


def remove_top_noise(input_file, noise_count=7):
    """Находит топ N самых частых слов и удаляет их из файла полностью."""
    with open(input_file, 'r', encoding='utf-8') as file:
        all_words = [line.strip() for line in file if line.strip()]

    # Находим, какие именно слова являются шумом
    word_counts = Counter(all_words)
    noise_words = set(word for word, count in word_counts.most_common(noise_count))
    # Когда делаем if word not in noise_words, Python в случае со списком каждый раз пробегает по нему.
    # В случае с set (множеством) поиск происходит мгновенно. На 600 словах разницы нет, но на 10 000+
    # это станет заметно.

    print(f"Удаляем шум: {', '.join(noise_words)}")

    # Оставляем только те слова, которых нет в списке шума
    filtered_words = [word for word in all_words if word not in noise_words]

    with open(input_file, 'w', encoding='utf-8') as f:
        for word in filtered_words:
            f.write(f"{word}\n")

    print(f"Файл {input_file} очищен от шума")
    return filtered_words  # Опционально


def remove_duplicates(input_file):
    """Читает файл и возвращает список уникальных слов в порядке их появления."""
    with open(input_file, 'r', encoding='utf-8') as file:
        words = [line.strip() for line in file if line.strip()]

    # Использование dict.fromkeys сохраняет порядок первого появления слова
    unique_words = list(dict.fromkeys(words))

    with open(input_file, 'w', encoding='utf-8') as f:
        for word in unique_words:
            f.write(f"{word}\n")

    print(f"Файл {input_file} успешно обновлен.")

    return unique_words  # Опционально


def translate_word(word, system_prompt, temperature=0.1, do_sample=False):
    """Генерирует перевод и справку для одного слова."""
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"Переведи слово: {word}"}
    ]

    text = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
    inputs = tokenizer([text], return_tensors="pt").to(model.device)

    # Генерируем ответ
    outputs = model.generate(**inputs, max_new_tokens=800, temperature=temperature, do_sample=do_sample)

    # Декодируем срезаем промпт, оставляем только ответ)
    response = tokenizer.decode(outputs[0][len(inputs.input_ids[0]):], skip_special_tokens=True)
    return response


def judge_adequacy(word, translation, judge_system_prompt):
    """Модель оценивает свою (или чужую) работу."""
    prompt = f"Word: {word}\nTranslation: {translation}\nRate adequacy (1-5):"

    messages = [
        {"role": "system", "content": judge_system_prompt},
        {"role": "user", "content": prompt}
    ]

    text = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
    inputs = tokenizer(text, return_tensors="pt").to(model.device)

    # Генерируем 2 токена (нам нужна только цифра)
    outputs = model.generate(**inputs, max_new_tokens=2, temperature=0.1, do_sample=False)

    score_raw = tokenizer.decode(outputs[0][len(inputs.input_ids[0]):], skip_special_tokens=True)

    match = re.search(r'\d', score_raw)
    return int(match.group()) if match else 0  # Возвращаем цифру или 0, если модель промолчала


def process_all_words(input_path, results_path, system_prompt, judge_system_prompt):
    """Читает список слов и сохраняет переводы по одному."""
    # Функция детекции китайских иероглифов
    def has_chinese(text):
        return any('\u4e00' <= char <= '\u9fff' for char in str(text))

    def has_unexpected_chars(text):
        # 1. Проверка на китайские иероглифы
        has_chinese = any('\u4e00' <= char <= '\u9fff' for char in str(text))

        # 2. Проверка на "мусорные" символы:
        # Ищем всё, что НЕ: латиница, кириллица, цифры, пробелы и пунктуация (. , ! ? - : ; ( ) [ ] " ')
        # Регулярка [^...] означает "всё, кроме перечисленного"
        bad_chars = re.findall(r'[^a-zA-Zа-яА-ЯёЁ0-9\s\.,!\?\-\:\;\(\)\[\]"\'«»—]', str(text))
        # Исключаем из "плохих" символов перенос строки, чтобы не ломать парсинг
        bad_chars = [c for c in bad_chars if c not in ['\n', '\r']]

        return has_chinese or len(bad_chars) > 0

    # Список для накопления строк таблицы
    rows = []
    # Список значимых колонок
    content_cols = ["Translation", "Examples", "Phrases"]

    with open(input_path, 'r', encoding='utf-8') as f:
        words_to_process = [line.strip() for line in f if line.strip()]

    print(f"Начинаю перевод {len(words_to_process)} слов в таблицу...")

    for word in words_to_process:
        try:
            max_attempts = 5
            attempt = 0
            success = False
            final_raw = ""
            final_score = 0

            # --- ЦИКЛ RETRY LOOP для рестарта перевода, если он неудовлетворителен ---
            while attempt < max_attempts and not success:
                attempt += 1

                # Первая попытка: железная строгость. Пересдача: добавляем "вариативность", чтобы выйти из тупика
                curr_temp = 0.1 if attempt == 1 else 0.7
                curr_sample = False if attempt == 1 else True

                current_prompt = system_prompt
                if attempt > 1:
                    # Усиливаем требования при повторной попытке
                    current_prompt  = system_prompt + f"\n\nCRITICAL ERROR: Your previous translation for '{word}' was REJECTED as a hallucination. " \
                                      f"DO NOT use non-existent words like 'себяг'. Use REAL Russian dictionary definitions." \
                                      f"All [T], [E], [P] sections must be present. NO Chinese. Presence of the target word in examples."

                # 1. Генерируем "сырой" ответ, с заменой температуры и сэмплами, если попыток > 1
                final_raw = translate_word(word, current_prompt, temperature=curr_temp, do_sample=curr_sample)

                # 2. ТЕХНИЧЕСКАЯ ПРОВЕРКА НАЛИЧИЯ ТЕГОВ (tags_present)
                tags_present = all(tag in final_raw for tag in ["[T]", "[E]", "[P]"])

                # 3. ВЫЗЫВАЕМ СУДЬЮ
                final_score = judge_adequacy(word, final_raw, judge_system_prompt)

                # 4. Проверяем на неожиданные символы
                unexpected_chars = has_unexpected_chars(final_raw)

                # УСЛОВИЕ УСПЕХА: Score >= 4 + все теги на месте + нет китайского
                fails = []
                if not tags_present: fails.append("пустые теги")
                if unexpected_chars: fails.append("неожиданные символы/иероглифы")
                if final_score < 4: fails.append(f"низкий score {final_score}")

                if not fails:  # Если список пуст — значит всё отлично
                    success = True
                    print(f"  [✓] {word}: Успех с попытки {attempt}")
                else:
                    # Соединяем все ошибки через запятую
                    reason = ", ".join(fails)
                    print(f"  [!] {word}: Попытка {attempt} не удалась ({reason}).")

            # # # --- РАЗБИВКА ТЕКСТА ---

            # # ПАРСИНГ
            # Ищем блоки с помощью "заглядывания вперед" (?=\[|$) :
            # Бери весь текст, пока не встретишь ЛИБО начало следующего тега [, ЛИБО самый конец сообщения $
            # Это позволяет найти текст, даже если это последний блок в ответе
            t_match = re.search(r'\[T\]\s*(.*?)\s*(?=\[|$)', final_raw, re.DOTALL)
            e_match = re.search(r'\[E\]\s*(.*?)\s*(?=\[|$)', final_raw, re.DOTALL)
            p_match = re.search(r'\[P\]\s*(.*?)\s*(?=\[|$)', final_raw, re.DOTALL)

            # Безопасно извлекаем текст. Если блок найден (.group(1)), чистим его (.strip()), если нет — пишем ошибку.
            res_t = t_match.group(1).strip() if t_match else "Parsing error"
            res_e = e_match.group(1).strip() if e_match else "Parsing error"
            res_p = p_match.group(1).strip() if p_match else "Parsing error"

            # Определяем финальный скор для таблицы
            if not success:
                status = "⚠ Галлюцинация" if final_score <= 2 else "⚠ Требует правки"
            else:
                status = "✓ OK"

            # # Если судья поставил 1-2, помечаем прямо в ячейке перевода
            # display_translation = f"[{final_score}/5] {res_t}" if final_score <= 2 else res_t

            # Записываем в итоговый список (добавляем новые колонки)
            rows.append({
                "English Word": word.upper(),
                "Translation": res_t,
                "Examples": res_e,
                "Phrases": res_p,
                "Score": final_score,  # колонка с цифрой
                "Status": status  # колонка с вердиктом
            })

            # Сохраняем
            df = pd.DataFrame(rows)
            df.to_excel(results_path, index=False,
                        engine='openpyxl')  # сохраняем отформатированные данные с помощью openpyxl

            # --- блок для красивого форматирования ---
            wb = load_workbook(results_path)
            ws = wb.active
            # TODO Если заметишь замедление, просто вынеси блок с load_workbook и Alignment за пределы цикла for, чтобы отформатировать всё один раз в самом конце. Но пока слов немного — оставляй внутри для надежности.

            # Задаем ширину колонок: Слово(15), Перевод(25), Остальное(50)
            dims = {'A': 25, 'B': 35, 'C': 75, 'D': 75, 'E': 10, 'F': 20}
            for col, value in dims.items():
                ws.column_dimensions[col].width = value

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

            wb.save(results_path)
            print(f"[✓] {word} добавлено в таблицу")

        except Exception as e:
            print(f"[!] Ошибка на слове {word}: {e}")

    # --- БЛОК МЕТРИК ---
    df = pd.DataFrame(rows)  # опционально. чтобы данные были доступны

    # Список колонок для проверки (чтобы Score и Status не портили статистику)
    content_cols = ["Translation", "Examples", "Phrases"]

    # Сбор статистики
    total_words = len(df)

    # Считаем ошибки парсинга только в контентных колонках
    parsing_errors_by_col = (df[content_cols] == "Parsing error").sum()
    total_parsing_errors = parsing_errors_by_col.sum()

    # Считаем пустые ячейки (используем .map вместо устаревшего .applymap)
    empty_cells = df[content_cols].map(lambda x: str(x).strip() == "").sum().sum()

    # Считаем иероглифы (хотя бы один в любой ячейке строки)
    chinese_count = df.apply(lambda row: row.map(has_chinese).any(), axis=1).sum()

    # --- ПЕЧАТЬ ОТЧЕТА ---
    print(f"\n" + "=" * 45)
    print(f"📊 ИТОГОВЫЙ ОТЧЕТ ПО КАЧЕСТВУ (Слов: {total_words})")
    print(f"⎯" * 45)

    print(f"❌ ТЕХНИЧЕСКИЙ БРАК:")
    for col in content_cols:
        print(f"  • {col}: {parsing_errors_by_col[col]} ошибок парсинга")
    print(f"  • Пустых ячеек (пропуски): {empty_cells}")

    if 'Score' in df.columns:
        hallucinations = (df['Score'] <= 2).sum()
        print(f"\n🧠 КАЧЕСТВО (LLM Judge):")
        print(f"  • Средняя адекватность: {df['Score'].mean():.2f} / 5.0")
        print(f"  • Галлюцинации (Score 1-2): {hallucinations}")

    print(f"\n🇨🇳 ЛОКАЛИЗАЦИЯ:")
    print(f"  • Слов с иероглифами: {chinese_count}")

    # Итоговый процент брака (только по контентным колонкам)
    total_content_cells = total_words * len(content_cols)
    if total_content_cells > 0:
        error_rate = ((total_parsing_errors + empty_cells) / total_content_cells) * 100
        print(f"⎯" * 45)
        print(f"📈 ОБЩИЙ ПРОЦЕНТ ТЕХНИЧЕСКОГО БРАКА: {error_rate:.2f}%")

    print("=" * 45)


# --- ЗАПУСК ---
if __name__ == "__main__":
    # """ Распознаем слова """
    # # Проверка существования папки перед запуском
    # if not os.path.exists(test_path):
    #     try:
    #         os.makedirs(test_path)
    #         print(f"Папка '{test_path}' создана. Пожалуйста, положите туда картинки.")
    #     except OSError:
    #         print(f"Не удалось создать папку {test_path}. Проверьте пути.")
    # # Формируем файл со словами
    # else:
    #     extractor = BatchEnglishExtractor()
    #     extractor.process_folder(test_path, OUTPUT_FILE)
    #
    # """ Подготовка таблицы уникальных слов """
    # remove_top_noise(OUTPUT_FILE, noise_count=7)
    # remove_duplicates(OUTPUT_FILE)
    # analyze_frequency(OUTPUT_FILE, top_n=20)

    """ Подключаем GPT 
    pip install huggingface_hub
    pip install bitsandbytes accelerate
    """

    print(f"CUDA Available: {torch.cuda.is_available()}")
    print(f"Current Device: {torch.cuda.get_device_name(0)}")
    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"[✓] Using device: {device}")

    model_name = "Qwen/Qwen2.5-7B-Instruct"  # Очень точные переводы и примеры

    # Создаем конфигурацию для 8-битного режима
    bnb_config = BitsAndBytesConfig(
        load_in_8bit=True
    )

    tokenizer = AutoTokenizer.from_pretrained(model_name, token=HF_TOKEN)  # TODO объяснить
    model = AutoModelForCausalLM.from_pretrained(
        model_name,
        #  torch_dtype=torch.float16,
        device_map="auto",
        quantization_config=bnb_config,  # если не хватает памяти. или load_in_4bit=True
        low_cpu_mem_usage=True,  # Загрузка по частям и  создание пустого каркаса
        token=HF_TOKEN
    )

    prompt_en_8_examples = """You are a translation engine. 
    RULES:
    1. ONLY English and Russian languages used.
    2. NO Chinese/Asian characters or translations.
    3. If you see Chinese in your thoughts, DELETE it.
    4. DO NOT use rare or non-exisent meanings.
    5. DO NOT make words up
    6. START each section strictly with the tags [T], [E], and [P].
    7. Each section tag must be filled. There must not be parsing errors
    8. Examples and phrases are well-composed and correctly translated to Russian
    9. [E] and [P]: Examples contain ONLY ENGLISH
    10. [E] and [P]: Example translations contain ONLY RUSSIAN
    11. [P]: If there are no well-known/set phrase with this word, it's allowed to leave "Not found"
    
    FORMAT:
    [T]
    (3-5 the most common Russian translations of this word, each on a new line)
    [E]
    (3 example phrases or full sentences using this word IN ENGLISH, each on a new line + hyphen + russian translation)
    [P]
    (1-3 set/well-known/slang phrases, or idioms (if they exist) using this word IN ENGLISH, each on a new line + hyphen + russian translation)

    EXAMPLE:
    Input: quizzically
    Output: 
    [T]
    насмешливо
    чудаковато
    с недоумением
    [E]
    She was sitting with her head quizzically tilted - она сидела с головой, наклоненной чудаковатым образом
    [P]
    With a quizzical expression - с недоумением"""

    judge_prompt_en_3 = """You are a ruthless Linguistic Auditor. 
    Your goal is to find errors in a dictionary entry. 

    CRITICAL ERRORS
    Score 1:
    - The Russian translation is a made-up word (hallucination like 'себяг').
    - The Russian translation is not found or empty
    - "Parsing error" in any section ([T], [E], or [P])
    
    Score 2:
    - Any Chinese/Asian characters found in any section ([T], [E], or [P]).
    - The Russian translation is completely unrelated to the English word or phrase.
    - Numbers found in any section
    - ([E] and [P] sections) the examples do not contain the word

    MINOR ERRORS (Score 3):
    - All 3 sections ([T], [E], or [P]) are present.
    - ([T] section) Only one translation provided instead of 3-5.
    - Russian grammar is broken.
    - The Russian translation is not the most common to the word
    - ([E] and [P] sections) Russian translation of phrases is incorrect
    
    GOOD (Score 4):
    - All 3 sections ([T], [E], or [P]) are present and accurate.
    - ONLY English and Russian languages used.
    - ([T] section) at least 2 accurate Russian words, corresponding the most common translation of the word.
    - ([E] section) at least 3 suitable well-written examples provided in english with correct Russian translation

    PERFECT (Score 5):
    - All 3 sections ([T], [E], or [P]) are present and accurate.
    - ONLY English and Russian languages used.
    - ([T] section) at least 3 accurate Russian words, corresponding the most common translation of the word.
    - ([E] section) at least 3 suitable well-written examples provided in english with correct Russian translation
    - ([P] section) at least 1 suitable phrase provided in english with correct Russian translation

    Respond ONLY with a single number from 1 to 5."""

    process_all_words(OUTPUT_FILE, EXCEL_FILE, prompt_en_8_examples, judge_prompt_en_3)
